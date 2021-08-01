import logging
import os.path
from pickle import STRING
import shutil
from pandas.core.frame import DataFrame

import win32com.client
import pandas as pd


from openpyxl import Workbook
from openpyxl import load_workbook
    
'''
Synchonize Outlook Tasks with an Excel file
Look at Readme.md for an overview of what this script does
'''
#Constants
LOG_FILE="task.log"
EXCEL_TASK_FILE="task-data.xlsx"
EXCEL_COL_NAMES={
    "Importance":1,
    "Role":2,
    "Categories":3,
    "Subject":4,
    "Team":5,
    "DueDate":6,
    "EntryID":7,
    "CreatedDate":8,
    "Modified":9
    }

#Handle TO Outlook, Logs and other objects we will need later
OUTLOOK = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")
logging.basicConfig(filename=LOG_FILE, encoding='utf-8', level=logging.DEBUG)

'''
Read Task from Excel, update into excel if possible
'''
def read_tasks_into_outlook():
    
    logging.info ("READING TASKS INTO OUTLOOK")

    # Read our Excel tasks into a pandas dataframe
    task_df_excel = pd.read_excel(EXCEL_TASK_FILE, index_col=False)
    task_df_excel.fillna('',inplace=True)
    task_df_excel.set_index('EntryID')

    logging.debug("Loaded from Excel")
    logging.debug(task_df_excel)


    #Now loop through our tasks in Outlook and try to match 
    thisFolder = OUTLOOK.GetDefaultFolder(13)
    folderItems = thisFolder.items

    for task in folderItems:
        this_task_id=str(task.EntryID)
        logging.debug("Looking for update to:"+this_task_id[-10:])

        # search our values from excel for this id
        if this_task_id not in task_df_excel.values:
            logging.debug("no match for:"+this_task_id[-10:]+ " skipping")

        else:

            logging.debug("found match for:"+this_task_id[-10:]+" processing")

            #set a filter for Pandas using the Task ID
            my_filter = task_df_excel['EntryID'] == this_task_id

            #Process this and decide to update
            rslt_df = task_df_excel.loc[my_filter]
            logging.debug("Matched in XL")
            logging.debug(rslt_df)

            #Now decide if we update or not
            modified_flag= rslt_df.iat[0,EXCEL_COL_NAMES["Modified"]-1]  #adjust for being index based
            if(modified_flag!='Y'):
                logging.debug("Modified flag not set to Y - ignoring")
            else:
                #do update
                logging.info("Modified is Y - try to update new text into task:"+rslt_df.iat[0,EXCEL_COL_NAMES["Subject"]-1])

                #Update the values
                #sheet.cell(row=2,column=EXCEL_COL_NAMES["Importance"]).value=task.Importance
                #sheet.cell(row=2,column=EXCEL_COL_NAMES["Role"]).value=task.Role 
                #sheet.cell(row=2,column=EXCEL_COL_NAMES["Categories"]).value=task.Categories # make comma safe?
                task.Subject= rslt_df.iat[0,EXCEL_COL_NAMES["Subject"]-1]
                print("attepting to update subject to:"+rslt_df.iat[0,EXCEL_COL_NAMES["Subject"]-1])
                print("now value:"+task.Subject)
                #sheet.cell(row=2,column=EXCEL_COL_NAMES["Team"]).value=task.TeamTask 
                
                #update Due Date only if it is not default
                #tmpDate = str(task.DueDate)
                #if(tmpDate!="4501-01-01 00:00:00+00:00"):
                #    sheet.cell(row=2,column=EXCEL_COL_NAMES["DueDate"]).value=tmpDate

                #sheet.cell(row=2,column=EXCEL_COL_NAMES["EntryID"]).value=task.EntryID 
                #sheet.cell(row=2,column=EXCEL_COL_NAMES["CreatedDate"]).value=str(task.CreationTime)
                #sheet.cell(row=2,column=EXCEL_COL_NAMES["Modified"]).value=str(task.LastModificationTime) 
        
    # ddd
    
    
 


'''
Clear the tasks output file, so we can reuse the formatting
'''
def clear_excel_output_file():
    logging.info ("CLEARING EXCEL TASK FILE")

    #Make a backup of the original file
    counter =1
    while(os.path.exists(str(counter)+EXCEL_TASK_FILE)):
        logging.debug("Backup file "+str(counter)+EXCEL_TASK_FILE+" exists, increment and try again")
        counter +=1

    shutil.copyfile(EXCEL_TASK_FILE, str(counter)+EXCEL_TASK_FILE)
    logging.debug("Created new backup file:"+str(counter)+EXCEL_TASK_FILE)

    #Open Sheet using Python
    workbook = load_workbook(filename=EXCEL_TASK_FILE)
    sheet = workbook.active

    #Now delete everything until we are only left with the header row
    # continuously delete row 2 until there
    # is only a single row left over 
    # that contains column names 
    while(sheet.max_row > 1):
        # this method removes the row 2
        logging.debug("deleting row")
        sheet.delete_rows(2)

    #Save the result
    workbook.save(filename=EXCEL_TASK_FILE)
    workbook.close


'''
Output Tasks from Outlook Into Excel
'''
def export_tasks_to_excel():
    thisFolder = OUTLOOK.GetDefaultFolder(13)

    folderItems = thisFolder.items
    logging.info ("EXPORTING TASKS TO EXCEL")
 
    #Open Excel Sheet using Python
    workbook = load_workbook(filename=EXCEL_TASK_FILE)
    sheet = workbook.active

    for task in folderItems:
        logging.debug("Outputting task:"+task.Subject)
        
        #insert a new clear line (shifting other tasks downwards)
        sheet.insert_rows(2)

        #Update the values
        sheet.cell(row=2,column=EXCEL_COL_NAMES["Importance"]).value=task.Importance
        sheet.cell(row=2,column=EXCEL_COL_NAMES["Role"]).value=task.Role 
        sheet.cell(row=2,column=EXCEL_COL_NAMES["Categories"]).value=task.Categories # make comma safe?
        sheet.cell(row=2,column=EXCEL_COL_NAMES["Subject"]).value=task.Subject # make comma safe?
        sheet.cell(row=2,column=EXCEL_COL_NAMES["Team"]).value=task.TeamTask 
        
        #update Due Date only if it is not default
        tmpDate = str(task.DueDate)
        if(tmpDate!="4501-01-01 00:00:00+00:00"):
            sheet.cell(row=2,column=EXCEL_COL_NAMES["DueDate"]).value=tmpDate

        sheet.cell(row=2,column=EXCEL_COL_NAMES["EntryID"]).value=task.EntryID 
        sheet.cell(row=2,column=EXCEL_COL_NAMES["CreatedDate"]).value=str(task.CreationTime)
        sheet.cell(row=2,column=EXCEL_COL_NAMES["Modified"]).value=str(task.LastModificationTime) 


    #Save the result
    workbook.save(filename=EXCEL_TASK_FILE)

# simple code to run from command line
if __name__ == '__main__':
    
    # Carry out the steps to sync excel adn outlook
    read_tasks_into_outlook()
    #clear_excel_output_file()
    #export_tasks_to_excel()
    
